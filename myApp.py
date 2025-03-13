import os
import glob
import sys
import random
import time
from colorama import *
from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill
from PyQt6.QtCore import QStringListModel, QAbstractItemModel
from PyQt6.QtGui import QStandardItemModel, QStandardItem
import pandas as pd
import random
#cài config selenium
from seleniumwire import webdriver  # Chặn request
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.alert import Alert
import openpyxl
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium_stealth import stealth
from urllib.parse import urlparse
import undetected_chromedriver as uc

from time import sleep
import PyQt6.uic as uic
from PyQt6.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog, QMessageBox
from myAppUi import Ui_Form  
class UI(QMainWindow):
    def __init__(self):
        super().__init__()
        
        # uic.loadUi("myApp.ui", self)
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        # self.bntChooseFile.clicked.connect(self.openFileDialog)
        # self.bntAction.clicked.connect(self.process)
        self.ui.bntChooseFile.clicked.connect(self.openFileDialog)
        self.ui.bntAction.clicked.connect(self.process)

        self.file_path = ""
            
    def openFileDialog(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Chọn file", "", "All Files (*);;Text Files (*.txt)")
        if file_path:
            self.ui.lbLinkfile.setText(f"Đã chọn: {file_path}")
            self.file_path = file_path
        else:
            self.ui.lbLinkfile.setText("Chưa chọn file")

    def process(self):
        if not self.file_path:
            QMessageBox.information(self, "Thông báo", "Chưa chọn file")
            return
        try:            
            df = pd.read_excel(self.file_path, engine="openpyxl", dtype=str)
            for index, row in df.iterrows():
                data = row.to_dict()
                IRSFormPage().action_form(data, index + 1, self.file_path)
                time.sleep(random.uniform(25, 50))
        except Exception as e:
            QMessageBox.information(self, "Thông báo", f"Có lỗi xảy ra: {str(e)}")
            
class IRSFormPage:
    def __init__(self):
        
        download_path = r"C:\Users\Admin\Downloads\IRS"  # Thư mục tải file
        # Cấu hình Selenium
        user_agents = [
            # Chrome trên Windows
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36",

            # Chrome trên macOS
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36",

            # Firefox trên Windows
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:92.0) Gecko/20100101 Firefox/92.0",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:93.0) Gecko/20100101 Firefox/93.0",

            # Firefox trên macOS
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:91.0) Gecko/20100101 Firefox/91.0",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:92.0) Gecko/20100101 Firefox/92.0",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:93.0) Gecko/20100101 Firefox/93.0",

            # Safari trên macOS
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.1 Safari/605.1.15",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.2 Safari/605.1.15",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.0 Safari/605.1.15",

            # Edge trên Windows
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Edg/91.0.864.59",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36 Edg/92.0.902.73",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36 Edg/93.0.961.52",

            # Edge trên macOS
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Edg/91.0.864.59",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36 Edg/92.0.902.73",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36 Edg/93.0.961.52",
        ]
        chrome_options  = uc.ChromeOptions()
        # Chọn ngẫu nhiên một User-Agent từ danh sách
        chrome_options.add_argument(f"user-agent={random.choice(user_agents)}")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")  # Ẩn chế độ tự động
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--allow-running-insecure-content")
        chrome_options.add_argument("--start-maximized")  # Mở rộng cửa sổ
        chrome_options.add_argument("--disable-gpu")    
        prefs = {
            "download.default_directory": download_path,  # Đặt thư mục tải file
            "download.prompt_for_download": False,          # Không hỏi xác nhận tải file
            "download.directory_upgrade": True,             # Cập nhật thư mục nếu cần
            "plugins.always_open_pdf_externally": True        # Nếu tải PDF, sẽ tải về thay vì mở trong trình duyệt
        }
        # Tắt extension tự động hóa
        # chrome_options.add_experimental_option("useAutomationExtension", False)
        chrome_options.add_experimental_option("prefs",prefs);
        self.driver = uc.Chrome(options=chrome_options)
        # self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        width = random.randint(800, 1200)
        height = random.randint(600, 800)
        self.driver.set_window_size(width, height)
        stealth(self.driver,
                languages=["en-US", "en"],
                vendor="Google Inc.",
                platform="Win32",
                webgl_vendor="Intel Inc.",
                renderer="Intel Iris OpenGL Engine",
                fix_hairline=True)
        
    def action_form(self, data, index, file_path):
        try:
            
            self.driver.delete_all_cookies() 
                      
            self.driver.get("https://sa.www4.irs.gov/modiein/individual/index.jsp")
            time.sleep(1)            #identifi Page
            try:
                alert = self.driver.switch_to.alert  # Chuyển sang alert
                print("Popup message:", alert.text)  # In ra nội dung popup
                alert.accept()  # Nhấn OK
                print("Đã nhấn OK trên popup.")
            except:
                print("Không tìm thấy popup.")

            
            self.click_button("//input[@type='submit' and @value='Begin Application >>']", "Begin Application")
        
            self.select_radio("//input[@type='radio' and @id='sole']")
                
            self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue")
    
            self.select_radio("//input[@type='radio' and @value='30' and @id='sole']")
                
            self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue")
        
            self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue")
                
            self.select_radio("//input[@type='radio' and @name='radioReasonForApplying']")
            
            self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue")
        
            self.auto_fill_form_Authenticate_2(data)
            self.auto_fill_form_Addresses_3(data)
            self.auto_fill_form_Details_4()
            self.auto_fill_form_EIN_Confirmation_5(index, file_path, data)
        except Exception as e:
            print(f"❌ Dừng xử lý dòng {index} do lỗi: {e}")
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            self.driver.delete_all_cookies()  # Xóa toàn bộ cookies
            # Xóa Local Storage & Session Storage
            self.driver.execute_script("window.localStorage.clear();")
            self.driver.execute_script("window.sessionStorage.clear();")

            self.driver.quit()
            return  # Thoát action_form(), quay lại vòng lặp tiếp theo
                   
    def auto_fill_form_Authenticate_2(self, data):
        self.fill_form_input("applicantFirstName", data["FN"])
        self.fill_form_input("applicantLastName", data["LN"])
        # Chuyển SSN thành chuỗi và đảm bảo đủ 9 số
        ssn_str = str(data["SSN"]).zfill(9)

        # Tách thành 3 phần
        ssn_part1 = ssn_str[:3]  # 3 số đầu
        ssn_part2 = ssn_str[3:5]  # 2 số giữa
        ssn_part3 = ssn_str[5:]   # 4 số cuối
        self.fill_form_input("applicantSSN3", ssn_part1)
        self.fill_form_input("applicantSSN2", ssn_part2)
        self.fill_form_input("applicantSSN4", ssn_part3)
        
        #click checkbox
        self.select_radio("//input[@type='radio' and @name='tpdQuestion' and @id='iamsole']")
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 
    
    def auto_fill_form_Addresses_3(self, data):
        self.fill_form_input("physicalAddressStreet", data["ADD"])
        self.fill_form_input("physicalAddressCity", data["CITY"])
        self.select_state(data["STATE"])
        self.fill_form_input("physicalAddressZipCode", data["ZIP"])
        
        first3 = str(random.randint(100, 999))  # 3 số đầu
        middle3 = str(random.randint(100, 999))  # 3 số giữa
        last4 = str(random.randint(1000, 9999))  # 4 số cuối
        
        self.fill_form_input("phoneFirst3", first3)
        self.fill_form_input("phoneMiddle3", middle3)
        self.fill_form_input("phoneLast4", last4)
        
        self.select_radio("//input[@type='radio' and @name='radioAnotherAddress' and @id='radioAnotherAddress_n' and @value='false']")
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 
        
    def auto_fill_form_Details_4(self):
        month_ramdom = str(random.randint(1,9))
        year_ramdom = str(random.randint(2022,2024))
        
        #page 1
        self.select_month(month_ramdom)
        self.fill_form_input("BUSINESS_OPERATIONAL_YEAR_ID", year_ramdom)
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 
        
        #page 2
        self.select_radio("//input[@type='radio' and @name='radioTrucking' and @id='radioTrucking_n' and @value='false']")
        self.select_radio("//input[@type='radio' and @name='radioInvolveGambling' and @id='radioInvolveGambling_n' and @value='false']")
        self.select_radio("//input[@type='radio' and @name='radioExciseTax' and @id='radioExciseTax_n' and @value='false']")
        self.select_radio("//input[@type='radio' and @name='radioSellTobacco' and @id='radioSellTobacco_n' and @value='false']")
        self.select_radio("//input[@type='radio' and @name='radioHasEmployees' and @id='radioHasEmployees_n' and @value='false']")
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 
            
        #page 3
        self.select_radio("//input[@type='radio' and @name='radioPrincipalActivity' and @id='other' and @value='15']")
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 
        
        #page 4
        self.select_radio("//input[@type='radio' and @name='radioPrincipalService' and @id='other' and @value='line15OpenEntry']")
        self.fill_form_input("pleasespecify", "digital")
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 
         #page 5
        #check ô đầu tiên
        self.select_radio("//input[@type='radio' and @name='radioLetterOption' and @id='receiveonline' and @value='online']")
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 
         #page 6
        self.click_button("//input[@type='submit' and @value='Submit'and @name='Submit']", "Submit completed") 
        
    def auto_fill_form_EIN_Confirmation_5(self, index, file_path, data):
        print("page 5, para, index excell  ", index,"  file path: ", file_path)
        ein_assigned = self.driver.find_element(By.XPATH, "//td[text()='EIN Assigned:']/following-sibling::td/b").text

        legal_name = self.driver.find_element(By.XPATH, "//td[text()='Legal Name:']/following-sibling::td/b").text

        self.save_to_excel(ein_assigned, legal_name, data)

                # Mở file Excel
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Ghi dữ liệu vào đúng dòng
        sheet.cell(row=index+1, column=1, value=ein_assigned)  # Cột 1 (EIN)
        sheet.cell(row=index+1, column=2, value=legal_name)  # C)
        
        # Lưu lại file Excel
        workbook.save(file_path)
        workbook.close()

        try:
            download_path = Path("C:/Users/Admin/Downloads/IRS").resolve()
            download_path.mkdir(parents=True, exist_ok=True)
            
            pdf_link_element = self.driver.find_element(By.XPATH, "//*[contains(@href, '.pdf')]")
            self.driver.execute_script("arguments[0].target='_self';", pdf_link_element)
            self.driver.execute_script("window.open = function() {};")

            pdf_url = pdf_link_element.get_attribute("href")
            print(f"📌 Link PDF: {pdf_url}")
            # Click để tải file
            # pdf_link_element.click()
            ActionChains(self.driver).key_down(Keys.CONTROL).click(pdf_link_element).key_up(Keys.CONTROL).perform()

            # Thời điểm click tải file
            download_start_time = time.time()

            timeout = 30
            start_time = time.time()
            latest_file = None

            while True:
                pdf_files = list(download_path.glob("*.pdf"))
                temp_files = list(download_path.glob("*.crdownload")) + list(download_path.glob("*.part"))

                # Lọc file PDF mới nhất sau khi tải
                if pdf_files:
                    latest_file = max(pdf_files, key=lambda f: f.stat().st_ctime)
                    if latest_file.stat().st_ctime > download_start_time and not temp_files:
                        break

                if time.time() - start_time > timeout:
                    print("❌ Lỗi: Quá thời gian chờ tải file")
                    latest_file = None
                    break

                time.sleep(2)         # Kiểm tra file có hợp lệ không
            if latest_file and latest_file.exists():
                new_filename = download_path / f"{legal_name}.pdf"

                # Nếu file trùng tên tồn tại, xóa trước khi đổi
                if new_filename.exists():
                    new_filename.unlink()

                latest_file.rename(new_filename)
                print(f"✅ File đã được đổi tên thành: {new_filename}")
            else:
                print("❌ Không tìm thấy file PDF để đổi tên")
                
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            self.driver.delete_all_cookies()  # Xóa toàn bộ cookies
            # Xóa Local Storage & Session Storage
            self.driver.execute_script("window.localStorage.clear();")
            self.driver.execute_script("window.sessionStorage.clear();")

            self.driver.quit()
        except Exception as e:
            print(f"❌ Lỗi khi tải PDF: {e}")
    
    def sleepRamdom(self, min_seconds=1, max_seconds=2):
        time.sleep(random.uniform(min_seconds, max_seconds))
      
    def save_to_excel(self, ein_assigned, legal_name, data):
        file_name = "export.xlsx"

        # Kiểm tra xem file đã tồn tại chưa
        if os.path.exists(file_name):
            wb = openpyxl.load_workbook(file_name)  # Mở file đã có
            ws = wb.active
        else:
            wb = openpyxl.Workbook()  # Tạo file mới
            ws = wb.active
            headers = ["ein_assigned", "legal_name", "SSN", "FN", "LN", "DOB", "ADD", "CITY", "STATE", "ZIP"]
            header_font = Font(bold=True, color="FF0000")
            header_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")

            # Ghi tiêu đề vào dòng 1
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        # Tìm dòng trống cuối cùng để ghi tiếp dữ liệu
        new_row = ws.max_row + 1

        # Ghi dữ liệu vào dòng mới
        ws.cell(row=new_row, column=1, value=ein_assigned)
        ws.cell(row=new_row, column=2, value=legal_name)
        ws.cell(row=new_row, column=3, value=data["SSN"])
        ws.cell(row=new_row, column=4, value=data["FN"])    
        ws.cell(row=new_row, column=5, value=data["LN"])    
        ws.cell(row=new_row, column=6, value=data["DOB"])    
        ws.cell(row=new_row, column=7, value=data["ADD"])    
        ws.cell(row=new_row, column=8, value=data["CITY"])    
        ws.cell(row=new_row, column=9, value=data["STATE"])    
        ws.cell(row=new_row, column=10, value=data["ZIP"])    

        # Lưu lại file
        wb.save(file_name)
        print(f"✅ Dữ liệu đã được thêm vào '{file_name}'!")


    
    def select_state(self, state):
        try:
            dropdown = Select(self.driver.find_element(By.ID, "physicalAddressState"))
            self.sleepRamdom()
            dropdown.select_by_value(state)
        except Exception as e:
            raise Exception(f"❌ Lỗi khi chọn state: {e}")
            
    def select_month(self, month):
        try:
            dropdown = Select(self.driver.find_element(By.ID, "BUSINESS_OPERATIONAL_MONTH_ID"))
            self.sleepRamdom()
            dropdown.select_by_value(month)
        except Exception as e:
            raise Exception(f"❌ Lỗi khi chọn state: {e}")
                
    def fill_form_input(self, xpath, data):
        try:
            WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, xpath))).send_keys(data)
            self.sleepRamdom()
            print(f"Đã điền {data} vào {xpath}.")
        except Exception as e:
            raise Exception(f"❌ Lỗi khi điền {data} vào {xpath}.")
     
    def select_radio(self, xpath):
        try:
            radio_button = WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath)))
            self.sleepRamdom()
            radio_button.click()
            print(f"Đã chọn radio.")
        except Exception as e:
            raise Exception(f"❌ Lỗi khi chọn radio: {e}")
    
    def click_button(self, xpath, description, repeat=1):
        for _ in range(repeat):
            try:
                button = WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                self.sleepRamdom()
                button.click()
                print(f"Đã nhấn nút {description}.")    
            except Exception as e:
                raise Exception(f"❌ Lỗi khi nhấn nút {description}: {e}")

                
            
if __name__ == "__main__":
    try:
        app = QApplication([])
        window = UI()
        window.show()
        app.exec()
    except KeyboardInterrupt:
        sys.exit()
            29