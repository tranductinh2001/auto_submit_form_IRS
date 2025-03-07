import os
import sys
import random
import time
import traceback
from colorama import *
from datetime import datetime
import json
import brotli
from PyQt6.QtCore import QStringListModel, QAbstractItemModel
from PyQt6.QtGui import QStandardItemModel, QStandardItem
import pandas as pd
import random
#cài config selenium
from seleniumwire import webdriver  # Chặn request
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.alert import Alert
import openpyxl

from time import sleep

from PyQt6 import uic
from PyQt6.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog, QMessageBox

class UI(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi("myApp.ui", self)
        self.bntChooseFile.clicked.connect(self.openFileDialog)
        self.bntAction.clicked.connect(self.process)
        
        self.file_path = ""
            
    def openFileDialog(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Chọn file", "", "All Files (*);;Text Files (*.txt)")
        if file_path:
            self.lbLinkfile.setText(f"Đã chọn: {file_path}")
            self.file_path = file_path
        else:
            self.lbLinkfile.setText("Chưa chọn file")

    def process(self):
        if not self.file_path:
            QMessageBox.information(self, "Thông báo", "Chưa chọn file")
            return
        try:            
            df = pd.read_excel(self.file_path, engine="openpyxl")

            # Hoặc hiển thị từng dòng
            for index, row in df.iterrows():
                data = row.to_dict()  
                IRSFormPage().action_form(data, index, self.file_path)
        except (OSError, IOError):
            QMessageBox.information(self, "Thông báo", "Có lỗi xảy ra")
            
class IRSFormPage:
    def __init__(self):
        chrome_options  = Options()
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")  # Ẩn chế độ tự động
        chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36")
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--allow-running-insecure-content")
        chrome_options.add_argument("--start-maximized")  # Mở rộng cửa sổ
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

    def action_form(self, data, index, file_path):
        # Truy cập trang web đăng nhập
        self.driver.get("https://sa.www4.irs.gov/modiein/individual/index.jsp")
        time.sleep(2)

        #identifi Page
        try:
            alert = self.driver.switch_to.alert  # Chuyển sang alert
            print("Popup message:", alert.text)  # In ra nội dung popup
            alert.accept()  # Nhấn OK
            print("Đã nhấn OK trên popup.")
        except:
            print("Không tìm thấy popup.")

        
        self.click_button("//input[@type='submit' and @value='Begin Application >>']", "Begin Application")
     
        self.select_radio("//input[@type='radio' and @id='sole']")
    
        self.select_radio("//input[@type='radio' and @value='30']")
            
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue")
 
        self.select_radio("//input[@type='radio' and @value='30' and @id='sole']")
            
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue")
    
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue")
            
        self.select_radio("//input[@type='radio' and @name='radioReasonForApplying']")
        
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue")
    
        self.auto_fill_form_Authenticate_2(data)
        self.auto_fill_form_Addresses_3(data)
        self.auto_fill_form_Details_4(data)
        # self.auto_fill_form_EIN_Confirmation_5(index, file_path)
        
    def auto_fill_form_Authenticate_2(self, data):
        self.fill_form_input("applicantFirstName", data["FN"])
        self.fill_form_input("applicantLastName", data["LN"])
        # Chuyển SSN thành chuỗi và đảm bảo đủ 9 số
        ssn_str = str(data["SSN"]).zfill(9)

        # Tách thành 3 phần
        ssn_part1 = ssn_str[:3]  # 3 số đầu
        ssn_part2 = ssn_str[3:5]  # 2 số giữa
        ssn_part3 = ssn_str[5:]   # 4 số cuối
        self.fill_form_input("applicantSSN3", ssn_part1)
        self.fill_form_input("applicantSSN2", ssn_part2)
        self.fill_form_input("applicantSSN4", ssn_part3)
        
        #click checkbox
        self.select_radio("//input[@type='radio' and @name='tpdQuestion' and @id='iamsole']")
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 
    
    def auto_fill_form_Addresses_3(self, data):
        self.fill_form_input("physicalAddressStreet", data["ADD"])
        self.fill_form_input("physicalAddressCity", data["CITY"])
        self.select_state(data["STATE"])
        self.fill_form_input("physicalAddressZipCode", data["ZIP"])
        
        first3 = str(random.randint(100, 999))  # 3 số đầu
        middle3 = str(random.randint(100, 999))  # 3 số giữa
        last4 = str(random.randint(1000, 9999))  # 4 số cuối
        
        self.fill_form_input("phoneFirst3", first3)
        self.fill_form_input("phoneMiddle3", middle3)
        self.fill_form_input("phoneLast4", last4)
        
        self.select_radio("//input[@type='radio' and @name='radioAnotherAddress' and @id='radioAnotherAddress_n' and @value='false']")
        
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 

    def auto_fill_form_Details_4(self, data):
        month_ramdom = str(random.randint(1,9))
        year_ramdom = str(random.randint(2022,2024))
        
        #page 1
        self.select_month(month_ramdom)
        self.fill_form_input("BUSINESS_OPERATIONAL_YEAR_ID", year_ramdom)
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 

        #page 2
        self.select_radio("//input[@type='radio' and @name='radioTrucking' and @id='radioTrucking_n' and @value='false']")
        self.select_radio("//input[@type='radio' and @name='radioInvolveGambling' and @id='radioInvolveGambling_n' and @value='false']")
        self.select_radio("//input[@type='radio' and @name='radioExciseTax' and @id='radioExciseTax_n' and @value='false']")
        self.select_radio("//input[@type='radio' and @name='radioSellTobacco' and @id='radioSellTobacco_n' and @value='false']")
        self.select_radio("//input[@type='radio' and @name='radioHasEmployees' and @id='radioHasEmployees_n' and @value='false']")
        
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 
            
        #page 3
        self.select_radio("//input[@type='radio' and @name='radioPrincipalActivity' and @id='other' and @value='15']")
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 
        
        #page 4
        self.select_radio("//input[@type='radio' and @name='radioPrincipalService' and @id='other' and @value='line15OpenEntry']")
        self.fill_form_input("pleasespecify", "digital")
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 

        #page 5
        #check ô đầu tiên
        self.select_radio("//input[@type='radio' and @name='radioLetterOption' and @id='receiveonline' and @value='online']")
        self.click_button("//input[@type='submit' and @value='Continue >>']", "Continue") 

        #page 6
        self.click_button("//input[@type='submit' and @value='Submit >>']", "Submit completed") 
        
    def auto_fill_form_EIN_Confirmation_5(self, index, file_path):
        # Tìm và lấy dữ liệu EIN Assigned
        ein = self.driver.find_element(By.XPATH, "//td[b[contains(text(), 'EIN Assigned')]]/following-sibling::td/b").text

        # Tìm và lấy dữ liệu Legal Name
        legal_name = self.driver.find_element(By.XPATH, "//td[b[contains(text(), 'Legal Name')]]/following-sibling::td/b").text

        # Mở file Excel
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Ghi dữ liệu vào đúng dòng
        sheet.cell(row=index, column=1, value=ein)  # Cột 1 (EIN)
        sheet.cell(row=index, column=2, value=legal_name)  # Cột 2 (Legal Name)
    
    def select_state(self, state):
        try:
            dropdown = Select(self.driver.find_element(By.ID, "physicalAddressState"))
            dropdown.select_by_value(state)
        except Exception as e:
            print(f"❌ Lỗi khi chọn state: {e}")
            
    def select_month(self, month):
        try:
            dropdown = Select(self.driver.find_element(By.ID, "BUSINESS_OPERATIONAL_MONTH_ID"))
            dropdown.select_by_value(month)
        except Exception as e:
            print(f"❌ Lỗi khi chọn state: {e}")
                
    def fill_form_input(self, xpath, data):
        try:
            WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, xpath))).send_keys(data)
            print(f"Đã điền {data['FN']} vào {xpath}.")
        except:
            print(f"Không tìm thấy {xpath}.")
     
    def select_radio(self, xpath):
        try:
            radio_button = WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath)))
            radio_button.click()
            print(f"Đã chọn radio.")
        except:
            print(f"Không tìm thấy radio.")
    
    def click_button(self, xpath, description, repeat=1):
        for _ in range(repeat):
            try:
                button = WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                button.click()
                print(f"Đã nhấn nút {description}.")
                time.sleep(2)
            except:
                print(f"Không tìm thấy nút {description}.")
                
                
            
if __name__ == "__main__":
    try:
        app = QApplication([])
        window = UI()
        window.show()
        app.exec()
    except KeyboardInterrupt:
        sys.exit()
            